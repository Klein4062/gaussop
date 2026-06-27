# -*- coding: utf-8 -*-
"""
字段规格审核表生成器 (v3: 去子项/单链流程 + 评审改进)
流程: 提出 → 评审 → 改进项分析 → 闭环 → 验收(谁提出谁验收,不可转单)
每张单据字段以五维呈现: 字段名 / 填写要求 / 字段类型 / 详细信息 / 特殊要求
填写要求取值: 必填 / 选填 / 条件必填 / 系统生成(用户免填) / 预填(可改)
输出: 质量改进诉求表单-字段规格审核.xlsx  (可用 Apple Numbers 直接打开审核)
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ---------------- 样式 ----------------
NAVY="1F3864"; BLUE="2E5496"; GRAY="F2F2F2"; WHITE="FFFFFF"
REQ="FCE4D6"; OPT="EDEDED"; COND="FFF2CC"; SYS="E2EFDA"; PREF="E4DFEC"

thin=Side(border_style="thin",color="B0B0B0")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

F_TITLE  =Font(name="微软雅黑",size=14,bold=True,color="FFFFFF")
F_SUB    =Font(name="微软雅黑",size=9,italic=True,color="44546A")
F_SECTION=Font(name="微软雅黑",size=11,bold=True,color="FFFFFF")
F_TBLH   =Font(name="微软雅黑",size=10,bold=True,color="FFFFFF")
F_NAME   =Font(name="微软雅黑",size=10,bold=True,color="333333")
F_VAL    =Font(name="微软雅黑",size=10,color="000000")
F_REQ    =Font(name="微软雅黑",size=10,bold=True,color="C00000")
F_OPT    =Font(name="微软雅黑",size=10,color="808080")
F_COND   =Font(name="微软雅黑",size=10,bold=True,color="806000")
F_SYS    =Font(name="微软雅黑",size=10,bold=True,color="375623")
F_PREF   =Font(name="微软雅黑",size=10,bold=True,color="5B2C6F")
F_NOTE   =Font(name="微软雅黑",size=9,italic=True,color="808080")

FILL_TITLE  =PatternFill("solid",fgColor=NAVY)
FILL_SECTION=PatternFill("solid",fgColor=BLUE)
FILL_TBLH   =PatternFill("solid",fgColor=BLUE)
FILL_NAME   =PatternFill("solid",fgColor=GRAY)
FILL_INPUT  =PatternFill("solid",fgColor=WHITE)

C =Alignment(horizontal="center",vertical="center",wrap_text=True)
L =Alignment(horizontal="left",  vertical="center",wrap_text=True,indent=1)
LT=Alignment(horizontal="left",  vertical="top",   wrap_text=True,indent=1)

HEADERS=["字段名","填写要求","字段类型","详细信息","特殊要求"]
WIDTHS =[26,14,16,40,36]

# 系统自动生成/继承的字段(用户免填)
SYS_NAMES={
 "主诉求单号","评审单号","分析单号","问题单号","需求单号","验收单号",
 "当前状态","关联主诉求单号","关联分析单号","关联闭环单号",
 "闭环类型",
 "提出日期","评审日期","分析日期","提出人(工号+姓名)","验收人",
 "序号","提交时间","提交人",
}

def eff_level(fname, level):
    return "系统生成" if fname in SYS_NAMES else level

def req_font(level):
    return {"必填":F_REQ,"选填":F_OPT,"条件必填":F_COND,"系统生成":F_SYS,"预填(可改)":F_PREF}.get(level,F_VAL)
def req_fill(level):
    return PatternFill("solid",fgColor={"必填":REQ,"选填":OPT,"条件必填":COND,"系统生成":SYS,"预填(可改)":PREF}.get(level,WHITE))


# ---------------- 字段数据 ----------------
FORMS=[]

# ① 提出
FORMS.append(("1-提出-质量改进诉求单",
    "主单号 QI-YYYY-NNN    ①提出 → ②评审 → ③改进项分析 → ④闭环 → ⑤验收",
    [
     ("一、主单信息",[
        ("主诉求单号","必填","编号","主单唯一编号","格式 QI-YYYY-NNN;全流程主键"),
        ("提出日期","必填","日期","单据创建日期","系统取提出时当前日期"),
        ("当前状态","必填","枚举(单选)","单据流转状态","初始默认『待评审』"),
     ]),
     ("二、人员信息",[
        ("提出人(工号+姓名)","必填","人员(工号+姓名)","提出人工号与姓名(成对)","系统取当前登录人"),
     ]),
     ("三、诉求总体内容",[
        ("诉求总体标题","必填","文本","一句话概括诉求",""),
        ("诉求来源","必填","枚举(单选)","诉求来源渠道","现网问题 / 客户提出 / 实验室发现"),
        ("关联yw系统单号","必填","关联","来源对应的yw系统单号","需关联来源业务系统的原始单号"),
        ("总体背景与目标","必填","长文本","问题概述、影响、期望达成",""),
     ]),
     ("四、审批",[
        ("审批人","必填","枚举(单选)","审批人(固定名单)","固定下拉名单,由系统配置;评审人=审批人"),
     ]),
    ]))

# ② 评审
FORMS.append(("2-评审-评审记录单",
    "关联主诉求单号    评审通过则进入分析并指派责任人;不通过则打回提出人修改",
    [
     ("一、评审信息",[
        ("评审单号","必填","编号","评审单唯一编号","系统生成"),
        ("关联主诉求单号","必填","关联","所属主诉求单","系统从主单继承"),
        ("评审日期","必填","日期","评审提交日期","系统取提交评审表单时当前日期"),
        ("评审人(工号+姓名)","必填","人员(工号+姓名)","评审人工号与姓名(成对)","系统继承提出单审批人(评审与审批为同一人)"),
     ]),
     ("二、评审结论",[
        ("评审结果","必填","枚举(单选)","本次评审结论","通过/不通过;不通过则打回提出人"),
        ("责任人(工号+姓名)","条件必填","人员(工号+姓名)","改进项责任人","评审结果为『通过』时必填;分析/闭环责任人由此继承"),
        ("不通过理由","条件必填","长文本","不通过的原因说明","评审结果为『不通过』时必填"),
     ]),
    ]))

# ③ 改进项分析
FORMS.append(("3-分析-改进项分析单",
    "关联:主诉求单号    接纳后填写闭环单号(校验合法性)进入闭环;不接纳则打回评审",
    [
     ("一、单据信息",[
        ("分析单号","必填","编号","分析单唯一编号","AN-YYYY-NNN;系统生成"),
        ("关联主诉求单号","必填","关联","所属主诉求单","系统继承"),
        ("分析日期","必填","日期","分析提交日期","系统取提交分析表单时当前日期"),
        ("当前状态","必填","枚举(单选)","单据状态","默认『分析中』"),
        ("责任人(工号+姓名)","必填","人员(工号+姓名)","改进项责任人","系统继承自评审阶段填写的责任人"),
     ]),
     ("二、分析与接纳",[
        ("接纳版本","条件必填","文本","本次接纳的版本号","是否接纳为『是』时必填(如 V1.0);不接纳则留空"),
        ("是否接纳","必填","枚举(单选)","是否接纳该改进项","是/否;否→打回评审阶段"),
        ("不接纳理由","条件必填","长文本","不接纳的原因说明","是否接纳为『否』时必填;接纳则留空"),
        ("闭环方法","条件必填","枚举(单选)","接纳后走哪条闭环","是否接纳为『是』时必填;问题单闭环/需求闭环"),
        ("计划完成期限","条件必填","日期","接纳后计划完成日期","是否接纳为『是』时必填"),
     ]),
     ("三、闭环单号(接纳后必填,需校验合法性)",[
        ("问题单号/需求单号","条件必填","编号","承接的闭环单号","是否接纳为『是』时必填;需校验编号合法性(格式正确且单据真实存在),校验通过方可进入闭环阶段"),
     ]),
     ("四、分析进展子项(可逐条单独提交,作为外部审视进展的依据)",[
        ("序号","必填","表格列","进展条目序号","系统自增"),
        ("进展说明","必填","长文本","本条进展内容描述","每条可单独提交"),
        ("提交时间","必填","日期","本条提交时间","系统取提交时当前时间"),
        ("提交人","必填","人员(工号+姓名)","提交人工号与姓名(成对)","系统取当前登录人"),
     ]),
    ]))

# ④-1 问题单闭环
FORMS.append(("4-1闭环-问题单闭环单",
    "关联:主诉求单号 + 分析单号    关注当前进展与闭环效果自测",
    [
     ("一、单据信息",[
        ("问题单号","必填","编号","问题闭环单唯一编号","PC-YYYY-NNN;系统生成"),
        ("关联主诉求单号","必填","关联","所属主诉求单","系统继承"),
        ("关联分析单号","必填","关联","承接分析单","系统继承"),
        ("当前状态","必填","枚举(单选)","单据状态","默认『实施中』"),
        ("闭环类型","必填","文本","闭环路径","固定『问题单闭环』"),
        ("责任人(工号+姓名)","必填","人员(工号+姓名)","改进项责任人","系统继承自分析阶段责任人"),
        ("SLA时间","必填","日期/时长","闭环完成时限","由闭环责任人填写并可修改"),
     ]),
     ("二、进展与自测",[
        ("当前进展","必填","长文本","当前实施进展描述",""),
        ("闭环效果自测","必填","富文本(可粘贴图片)","自测的闭环效果","支持富文本与粘贴图片"),
     ]),
     ("三、进展子项(可逐条单独提交,作为外部审视进展的依据)",[
        ("序号","必填","表格列","进展条目序号","系统自增"),
        ("进展说明","必填","长文本","本条进展内容描述","每条可单独提交"),
        ("提交时间","必填","日期","本条提交时间","系统取提交时当前时间"),
        ("提交人","必填","人员(工号+姓名)","提交人工号与姓名(成对)","系统取当前登录人"),
     ]),
    ]))

# ④-2 需求闭环
FORMS.append(("4-2闭环-需求闭环单",
    "关联:主诉求单号 + 分析单号    关注当前进展与闭环效果自测",
    [
     ("一、单据信息",[
        ("需求单号","必填","编号","需求闭环单唯一编号","RC-YYYY-NNN;系统生成"),
        ("关联主诉求单号","必填","关联","所属主诉求单","系统继承"),
        ("关联分析单号","必填","关联","承接分析单","系统继承"),
        ("当前状态","必填","枚举(单选)","单据状态","默认『实施中』"),
        ("闭环类型","必填","文本","闭环路径","固定『需求闭环』"),
        ("责任人(工号+姓名)","必填","人员(工号+姓名)","改进项责任人","系统继承自分析阶段责任人"),
        ("SLA时间","必填","日期/时长","闭环完成时限","由闭环责任人填写并可修改"),
     ]),
     ("二、进展与自测",[
        ("当前进展","必填","长文本","当前实施进展描述",""),
        ("闭环效果自测","必填","富文本(可粘贴图片)","自测的闭环效果","支持富文本与粘贴图片"),
     ]),
     ("三、进展子项(可逐条单独提交,作为外部审视进展的依据)",[
        ("序号","必填","表格列","进展条目序号","系统自增"),
        ("进展说明","必填","长文本","本条进展内容描述","每条可单独提交"),
        ("提交时间","必填","日期","本条提交时间","系统取提交时当前时间"),
        ("提交人","必填","人员(工号+姓名)","提交人工号与姓名(成对)","系统取当前登录人"),
     ]),
    ]))

# ⑤ 验收(谁提出谁验收,不可转单)
FORMS.append(("5-验收-验收单",
    "关联:主诉求单号 + 闭环单号    验收人=改进提出人,不可转单",
    [
     ("一、单据信息",[
        ("验收单号","必填","编号","验收单唯一编号","格式 AC-YYYY-NNN"),
        ("关联主诉求单号","必填","关联","所属主诉求单","系统继承"),
        ("关联闭环单号","必填","关联","承接的闭环单","系统继承"),
        ("验收人","必填","人员(工号+姓名)","验收人工号与姓名(成对)","=改进提出人,系统生成,锁定不可转单"),
        ("当前状态","必填","枚举(单选)","单据状态","默认『待验收』"),
     ]),
     ("二、验收结论",[
        ("验收结论","必填","富文本(可粘贴图片)","验收结论说明","支持富文本与粘贴图片"),
        ("验收是否通过","必填","枚举(单选)","是否通过验收","通过/不通过;通过即完成"),
     ]),
    ]))


# ---------------- 渲染 ----------------
wb=openpyxl.Workbook()

ws0=wb.active; ws0.title="0-阅读说明"; ws0.sheet_view.showGridlines=False
ws0.column_dimensions["A"].width=3
ws0.column_dimensions["B"].width=26
ws0.column_dimensions["C"].width=80
ws0.merge_cells("B1:C1")
c=ws0.cell(1,2,"字段规格审核表 · 阅读说明"); c.fill=FILL_TITLE; c.font=F_TITLE; c.alignment=C
ws0.row_dimensions[1].height=30

legend=[
 ("字段维度","字段名 / 填写要求 / 字段类型 / 详细信息 / 特殊要求  共 5 列"),
 ("填写要求 · 必填","红字橙底 —— 出单时必须填写"),
 ("填写要求 · 选填","灰字 —— 可不填"),
 ("填写要求 · 条件必填","黄字黄底 —— 满足某条件时必填(详见『特殊要求』)"),
 ("填写要求 · 系统生成","绿字绿底 —— 系统自动生成/继承,用户免填(编号/状态/关联/日期/责任人继承/提交人等)"),
 ("填写要求 · 预填(可改)","紫字紫底 —— 系统预填默认值,用户可修改"),
 ("字段类型 · 编号","单据唯一编号,有固定格式(如 QI-YYYY-NNN)"),
 ("字段类型 · 关联","引用其他单据/系统的编号,多由系统继承(如关联yw系统单号需用户填写)"),
 ("字段类型 · 文本 / 长文本","简短文字 / 多行描述"),
 ("字段类型 · 日期 / 日期·时长","日期(YYYY-MM-DD);SLA时间可为日期或时长"),
 ("字段类型 · 枚举(单选)","下拉单选,可选值固定(审批人为固定名单)"),
 ("字段类型 · 富文本(可粘贴图片)","支持富文本与粘贴图片(闭环自测/验收结论)"),
 ("字段类型 · 人员(工号+姓名)","工号与姓名成对录入"),
 ("字段类型 · 表格列","进展子项表中的列"),
]
r=3
ws0.cell(r,2,"图例").font=F_SECTION
r+=1
for k,v in legend:
    ws0.merge_cells(start_row=r,start_column=2,end_row=r,end_column=2)
    a=ws0.cell(r,2,k); a.fill=FILL_NAME; a.font=F_NAME; a.alignment=L; a.border=BORDER
    ws0.merge_cells(start_row=r,start_column=3,end_row=r,end_column=3)
    b=ws0.cell(r,3,v); b.font=F_VAL; b.alignment=L; b.border=BORDER
    ws0.row_dimensions[r].height=24; r+=1

r+=1
ws0.merge_cells(start_row=r,start_column=2,end_row=r,end_column=3)
n=ws0.cell(r,2,"单链流程(不拆子项):提出 → 评审 → 改进项分析 → 闭环 → 验收。"
              "关键规则:①评审完成前提出单可编辑;②评审不通过打回提出、分析不接纳打回评审;"
              "③责任人 评审填写→分析继承→闭环继承;④谁提出谁验收,不可转单。")
n.font=F_NOTE; n.alignment=L; ws0.row_dimensions[r].height=56


def render_form(name,subtitle,sections):
    ws=wb.create_sheet(name); ws.sheet_view.showGridlines=False
    for i,w in enumerate(WIDTHS,start=1):
        ws.column_dimensions[get_column_letter(i)].width=w
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=5)
    t=ws.cell(1,1,name); t.fill=FILL_TITLE; t.font=F_TITLE; t.alignment=C
    ws.row_dimensions[1].height=30
    ws.merge_cells(start_row=2,start_column=1,end_row=2,end_column=5)
    s=ws.cell(2,1,subtitle); s.font=F_SUB; s.alignment=L
    ws.row_dimensions[2].height=18
    r=3
    for i,h in enumerate(HEADERS,start=1):
        c=ws.cell(r,i,h); c.fill=FILL_TBLH; c.font=F_TBLH; c.alignment=C; c.border=BORDER
    ws.row_dimensions[r].height=22; r+=1
    for sec_title,fields in sections:
        ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=5)
        sc=ws.cell(r,1,sec_title); sc.fill=FILL_SECTION; sc.font=F_SECTION; sc.alignment=L
        for i in range(1,6): ws.cell(r,i).border=BORDER
        ws.row_dimensions[r].height=22; r+=1
        for (fname,level,ftype,detail,special) in fields:
            level=eff_level(fname,level)
            a=ws.cell(r,1,fname); a.font=F_NAME; a.fill=FILL_NAME; a.alignment=L; a.border=BORDER
            b=ws.cell(r,2,level); b.font=req_font(level); b.fill=req_fill(level); b.alignment=C; b.border=BORDER
            c=ws.cell(r,3,ftype); c.font=F_VAL; c.alignment=C; c.border=BORDER
            d=ws.cell(r,4,detail or "—"); d.font=F_VAL; d.alignment=LT; d.border=BORDER
            e=ws.cell(r,5,special or "—"); e.font=F_VAL; e.alignment=LT; e.border=BORDER
            ws.row_dimensions[r].height=max(22, 14*max(1, len(detail)//18+1, len(special)//16+1))
            r+=1
    return ws

for name,subtitle,sections in FORMS:
    render_form(name,subtitle,sections)

for ws in wb.worksheets:
    if ws.title!="0-阅读说明":
        ws.freeze_panes="A4"

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"质量改进诉求表单-字段规格审核.xlsx")
wb.save(out)
print("saved:",out)
print("sheets:",wb.sheetnames)
for name,subtitle,sections in FORMS:
    fields=[(x[0],eff_level(x[0],x[1])) for _,f in sections for x in f]
    n=len(fields); lvl={}
    for _,lv in fields: lvl[lv]=lvl.get(lv,0)+1
    print(f"  {name}: {n} 字段  " + "  ".join(f"{k}{v}" for k,v in lvl.items()))
