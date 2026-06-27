# -*- coding: utf-8 -*-
"""
质量改进诉求流程表单生成器 (v4: 单链流程 + 评审改进)
流程: 提出 → 评审 → 改进项分析 → 闭环(问题单/需求) → 验收(谁提出谁验收,不可转单)
关键规则: 评审完成前提出单可编辑;评审不通过打回提出、分析不接纳打回评审;
         责任人 评审填写→分析继承→闭环继承;闭环自测/验收结论支持富文本与图片。
填写类型视觉约定: 用户填写(白) / 系统生成(绿,带『系统生成』) / 预填可改(紫,带『预填·可改』)
输出: 质量改进诉求流程表单.xlsx
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

# ---------------- 样式 ----------------
NAVY="1F3864"; BLUE="2E5496"; STEEL="8EAADB"; LBLUE="D9E1F2"
GRAY="F2F2F2"; CREAM="FFF8E1"; WHITE="FFFFFF"; SUBROW="FBF6E2"
SYS_C="E2EFDA"; PRE_C="E4DFEC"

thin=Side(border_style="thin",color="B0B0B0")
BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)

F_TITLE  =Font(name="微软雅黑",size=15,bold=True,color="FFFFFF")
F_SUB    =Font(name="微软雅黑",size=9,italic=True,color="44546A")
F_SECTION=Font(name="微软雅黑",size=11,bold=True,color="FFFFFF")
F_LABEL  =Font(name="微软雅黑",size=10,bold=True,color="333333")
F_VALUE  =Font(name="微软雅黑",size=10,color="000000")
F_HINT   =Font(name="微软雅黑",size=9,italic=True,color="707070")
F_NOTE   =Font(name="微软雅黑",size=9,italic=True,color="808080")
F_BOX    =Font(name="微软雅黑",size=10,bold=True,color="FFFFFF")
F_TBLH   =Font(name="微软雅黑",size=10,bold=True,color="FFFFFF")
F_SAMP   =Font(name="微软雅黑",size=9,color="595959")

FILL_TITLE  =PatternFill("solid",fgColor=NAVY)
FILL_SECTION=PatternFill("solid",fgColor=BLUE)
FILL_TBLH  =PatternFill("solid",fgColor=BLUE)
FILL_BOX    =PatternFill("solid",fgColor=STEEL)
FILL_LABEL  =PatternFill("solid",fgColor=GRAY)
FILL_INPUT  =PatternFill("solid",fgColor=WHITE)
FILL_LONG   =PatternFill("solid",fgColor=CREAM)
FILL_SUB    =PatternFill("solid",fgColor=LBLUE)
FILL_SAMP   =PatternFill("solid",fgColor=SUBROW)
FILL_SYS    =PatternFill("solid",fgColor=SYS_C)
FILL_PRE    =PatternFill("solid",fgColor=PRE_C)

C =Alignment(horizontal="center",vertical="center",wrap_text=True)
L =Alignment(horizontal="left",  vertical="center",wrap_text=True,indent=1)
LT=Alignment(horizontal="left",  vertical="top",   wrap_text=True,indent=1)

USER=None; SYS="sys"; PRE="pre"

def box(ws,r1,c1,r2,c2):
    for r in range(r1,r2+1):
        for c in range(c1,c2+1):
            ws.cell(r,c).border=BORDER

def merge_val(ws,r,c1,c2,value="",fill=FILL_INPUT,font=F_VALUE,align=L):
    ws.merge_cells(start_row=r,start_column=c1,end_row=r,end_column=c2)
    cell=ws.cell(r,c1,value); cell.fill=fill; cell.font=font; cell.alignment=align
    return cell

def style_input(cell, mode):
    if mode==SYS:
        cell.fill=FILL_SYS; cell.value="(系统生成)"; cell.font=F_HINT; cell.alignment=LT
    elif mode==PRE:
        cell.fill=FILL_PRE; cell.value="(预填·可改)"; cell.font=F_HINT; cell.alignment=LT
    else:
        cell.fill=FILL_INPUT; cell.font=F_VALUE; cell.alignment=LT
    return cell


# ---------------- 枚举 ----------------
OPT={
 "来源":["现网问题","客户提出","实验室发现"],
 "状态":["待评审","评审中","分析中","实施中","待验收","已完成","已驳回"],
 "评审结果":["通过","不通过"],
 "是否":["是","否"],
 "闭环方法":["问题单闭环","需求闭环"],
 "验收是否通过":["通过","不通过"],
 "审批人":["张三(10001)","李四(10002)","王五(10003)"],   # 固定下拉名单(示例,实际由系统配置)
}


# ---------------- 表单渲染器 ----------------
def render_form(ws,title,subtitle,sections):
    ws.sheet_view.showGridlines=False
    for c,w in {"A":17,"B":15,"C":15,"D":17,"E":15,"F":15}.items():
        ws.column_dimensions[c].width=w

    def add_dropdown(coord,options):
        if not options: return
        dv=DataValidation(type="list",formula1='"'+",".join(options)+'"',allow_blank=True)
        ws.add_data_validation(dv); dv.add(coord)

    def parse(f):
        if not isinstance(f,tuple): return (f,None,USER)
        f=list(f)+[None]*(3-len(f))
        return (f[0], f[1], f[2] or USER)

    r=1
    merge_val(ws,r,1,6,title,FILL_TITLE,F_TITLE,C); ws.row_dimensions[r].height=32; r+=1
    if subtitle:
        merge_val(ws,r,1,6,subtitle,FILL_SUB,F_SUB,C); ws.row_dimensions[r].height=16; r+=1

    for sec in sections:
        merge_val(ws,r,1,6,sec["title"],FILL_SECTION,F_SECTION,L); box(ws,r,1,r,6)
        ws.row_dimensions[r].height=22; r+=1
        for rowdef in sec["fields"]:
            kind=rowdef[0]
            if kind=="SINGLE":
                label=rowdef[1]; height=rowdef[2] if len(rowdef)>2 else 36
                options=rowdef[3] if len(rowdef)>3 else None
                mode  =rowdef[4] if len(rowdef)>4 else USER
                lc=ws.cell(r,1,label); lc.font=F_LABEL; lc.fill=FILL_LABEL; lc.alignment=C
                cell=merge_val(ws,r,2,6); style_input(cell,mode); box(ws,r,1,r,6)
                if mode!=SYS: add_dropdown(ws.cell(r,2).coordinate,options)
                ws.row_dimensions[r].height=height; r+=1
            elif kind=="TWO":
                l1,o1,m1=parse(rowdef[1]); l2,o2,m2=parse(rowdef[2])
                cc=ws.cell(r,1,l1); cc.font=F_LABEL; cc.fill=FILL_LABEL; cc.alignment=C
                c1=merge_val(ws,r,2,3); style_input(c1,m1)
                cc=ws.cell(r,4,l2); cc.font=F_LABEL; cc.fill=FILL_LABEL; cc.alignment=C
                c2=merge_val(ws,r,5,6); style_input(c2,m2); box(ws,r,1,r,6)
                if m1!=SYS: add_dropdown(ws.cell(r,2).coordinate,o1)
                if m2!=SYS: add_dropdown(ws.cell(r,5).coordinate,o2)
                ws.row_dimensions[r].height=24; r+=1
            elif kind=="TABLE":
                headers=rowdef[1]; col_spans=rowdef[2]; n_rows=rowdef[3]
                col_modes=rowdef[4] if len(rowdef)>4 else [USER]*len(headers)
                col_opts =rowdef[5] if len(rowdef)>5 else [None]*len(headers)
                for (a,b),h in zip(col_spans,headers):
                    ws.merge_cells(start_row=r,start_column=a,end_row=r,end_column=b)
                    cell=ws.cell(r,a,h); cell.fill=FILL_TBLH; cell.font=F_TBLH; cell.alignment=C
                    box(ws,r,a,r,b)
                ws.row_dimensions[r].height=20; r+=1
                for _ in range(n_rows):
                    for (a,b),mode,opt in zip(col_spans,col_modes,col_opts):
                        ws.merge_cells(start_row=r,start_column=a,end_row=r,end_column=b)
                        cell=ws.cell(r,a); style_input(cell,mode); box(ws,r,a,r,b)
                        if mode!=SYS: add_dropdown(ws.cell(r,a).coordinate,opt)
                    ws.row_dimensions[r].height=26; r+=1
        r+=1
    return r


# =====================================================================
wb=openpyxl.Workbook()

# ---------------- Sheet 0: 流程总览 ----------------
ws=wb.active; ws.title="0-流程总览"; ws.sheet_view.showGridlines=False
for c,w in {"A":17,"B":15,"C":15,"D":17,"E":15,"F":15}.items():
    ws.column_dimensions[c].width=w
merge_val(ws,1,1,6,"质量改进诉求处理流程（提出 → 评审 → 改进项分析 → 闭环 → 验收）",FILL_TITLE,F_TITLE,C)
ws.row_dimensions[1].height=34
merge_val(ws,3,1,6,
    "① 提出  ➜  ② 评审(通过→指派责任人;不通过→打回提出)  ➜  ③ 改进项分析(接纳→定闭环方法→挂闭环单号;不接纳→打回评审)  ➜  ④ 闭环(问题单/需求)  ➜  ⑤ 验收(谁提出谁验收,通过即完成)",
    FILL_BOX,F_BOX,C); box(ws,3,1,3,6); ws.row_dimensions[3].height=40
merge_val(ws,4,1,6,"单链流程（不拆子项）；④ 闭环含两条路径：问题单闭环（缺陷/故障纠正预防）｜ 需求闭环（新功能/优化上线）",FILL_SUB,F_NOTE,C)
box(ws,4,1,4,6); ws.row_dimensions[4].height=18

merge_val(ws,6,1,6,"流转规则",FILL_SECTION,F_SECTION,L); box(ws,6,1,6,6); ws.row_dimensions[6].height=20
rules=[
 "1. 一张【主诉求单】即一项改进措施（不拆子项），整链流转：提出 → 评审 → 改进项分析 → 闭环 → 验收。",
 "2. 评审完成前，提出单可由提出人编辑修改。",
 "3. 评审通过则进入分析并由评审指派【责任人】；不通过则打回提出人修改。",
 "4. 改进项分析：接纳后决定闭环方法（问题单/需求）并挂闭环单号（校验合法性）后进入闭环；不接纳则打回评审。",
 "5. 责任人链：评审填写 → 分析继承 → 闭环继承。",
 "6. 验收：谁提出谁验收，验收人=提出人，锁定不可转单；通过即完成。",
 "7. 闭环自测/验收结论支持富文本与粘贴图片；分析与闭环均有可逐条单独提交的【进展子项】，供外部审视。",
]
r=7
for t in rules:
    merge_val(ws,r,1,6,t,FILL_INPUT,F_VALUE,L); box(ws,r,1,r,6); ws.row_dimensions[r].height=26; r+=1

r+=1
hh=["阶段","对应表单(Sheet)","目的","关键判定"]
colspans=[(1,1),(2,2),(3,5),(6,6)]
for (a,b),htxt in zip(colspans,hh):
    ws.merge_cells(start_row=r,start_column=a,end_row=r,end_column=b)
    cell=ws.cell(r,a,htxt); cell.fill=FILL_TBLH; cell.font=F_TBLH; cell.alignment=C
box(ws,r,1,r,6); ws.row_dimensions[r].height=22; r+=1
rows=[
 ("① 提出","1-提出-质量改进诉求单","记录诉求(关联yw系统单号)","是否受理"),
 ("② 评审","2-评审-评审记录单","评审、指派责任人","通过/不通过(打回提出)"),
 ("③ 分析","3-分析-改进项分析单","接纳与否、定闭环方法、挂闭环单号","是否接纳(不接纳打回评审)"),
 ("④a 闭环","4-1闭环-问题单闭环单","问题单:进展与效果自测","进展/自测/SLA"),
 ("④b 闭环","4-2闭环-需求闭环单","需求:进展与效果自测","进展/自测/SLA"),
 ("⑤ 验收","5-验收-验收单","谁提出谁验收,通过即完成","验收是否通过"),
 ("★ 跟踪","1.5-改进诉求台账(中枢)","汇总各阶段进度","是否完成"),
]
for row in rows:
    for (a,b),val in zip(colspans,row):
        ws.merge_cells(start_row=r,start_column=a,end_row=r,end_column=b)
        cell=ws.cell(r,a,val); cell.font=F_VALUE; cell.alignment=LT
        if a==1: cell.fill=FILL_LABEL; cell.font=F_LABEL; cell.alignment=C
    box(ws,r,1,r,6); ws.row_dimensions[r].height=30; r+=1


# ---------------- Sheet 1: 提出 ----------------
ws1=wb.create_sheet("1-提出-质量改进诉求单")
render_form(ws1,"质量改进诉求单（提出阶段）",
    "主单号 QI-YYYY-NNN    ①提出 → ②评审 → ③改进项分析 → ④闭环 → ⑤验收",[
    {"title":"一、主单信息","fields":[
        ("TWO",("主诉求单号",None,SYS),("提出日期",None,SYS)),
        ("SINGLE","当前状态",22,OPT["状态"],SYS),
    ]},
    {"title":"二、人员信息","fields":[
        ("SINGLE","提出人(工号+姓名)",24,None,SYS),
    ]},
    {"title":"三、诉求总体内容","fields":[
        ("SINGLE","诉求总体标题",26),
        ("TWO",("诉求来源",OPT["来源"]),"关联yw系统单号"),
        ("SINGLE","总体背景与目标（问题概述、影响、期望达成）",50),
    ]},
    {"title":"四、审批","fields":[
        ("SINGLE","审批人(固定下拉名单)",24,OPT["审批人"]),
    ]},
])


# ---------------- Sheet 2: 评审 ----------------
ws2=wb.create_sheet("2-评审-评审记录单")
render_form(ws2,"评审记录单（评审阶段 · 通过则进入分析并指派责任人）",
    "关联主诉求单号    评审通过则进入分析并指派责任人;不通过则打回提出人修改",[
    {"title":"一、评审信息","fields":[
        ("TWO",("评审单号",None,SYS),("关联主诉求单号",None,SYS)),
        ("TWO",("评审日期",None,SYS),("评审人(工号+姓名)",None,SYS)),
    ]},
    {"title":"二、评审结论","fields":[
        ("TWO",("评审结果",OPT["评审结果"]),"责任人(工号+姓名)"),
        ("SINGLE","不通过理由（评审结果为『不通过』时必填）",40),
    ]},
])


# ---------------- Sheet 1.5: 改进诉求台账 ----------------
wsL=wb.create_sheet("1.5-改进诉求台账(中枢)")
wsL.sheet_view.showGridlines=False
heads=["主诉求单号","提出人","诉求来源","关联yw系统单号","评审结果","责任人",
       "分析·是否接纳","闭环方法","闭环单号","SLA时间","验收·是否通过","当前状态"]
widths=[15,12,11,15,10,12,12,12,13,12,13,11]
for i,w in enumerate(widths,start=1):
    wsL.column_dimensions[get_column_letter(i)].width=w
merge_val(wsL,1,1,12,"改进诉求台账（中枢 · 汇总各阶段进度）",FILL_TITLE,F_TITLE,C)
wsL.row_dimensions[1].height=32
merge_val(wsL,2,1,12,"一张主诉求单即一项改进措施,单链流转;本表汇总其在各阶段的结果与状态。"
                     "『验收·是否通过』为通过即完成。",FILL_SUB,F_SUB,L)
wsL.row_dimensions[2].height=30
r=3
for i,h in enumerate(heads,start=1):
    cell=wsL.cell(r,i,h); cell.fill=FILL_TBLH; cell.font=F_TBLH; cell.alignment=C; cell.border=BORDER
wsL.row_dimensions[r].height=26
samples=[
 ["QI-2026-001","张三/工艺部","现网问题","YW-2026-007","通过","张工","是","问题单闭环","PC-2026-001","2026-07-10","通过","已完成"],
 ["QI-2026-002","李四/研发部","客户提出","YW-2026-019","通过","李工","是","需求闭环","RC-2026-002","2026-07-20","待验收","待验收"],
 ["QI-2026-003","王五/质量部","实验室发现","YW-2026-031","不通过","—","—","—","—","—","—","已驳回"],
]
r=4
for row in samples:
    for i,v in enumerate(row,start=1):
        cell=wsL.cell(r,i,v); cell.fill=FILL_SAMP; cell.font=F_SAMP
        cell.alignment=C if i!=2 else LT; cell.border=BORDER
    wsL.row_dimensions[r].height=22; r+=1
for _ in range(8):
    for i in range(1,13):
        cell=wsL.cell(r,i); cell.fill=FILL_INPUT; cell.border=BORDER; cell.alignment=LT
    wsL.row_dimensions[r].height=22; r+=1
merge_val(wsL,r+1,1,12,"注：浅色行为示例数据,实际使用时替换或删除。",FILL_SUB,F_NOTE,L)


# ---------------- Sheet 3: 改进项分析 ----------------
ws3=wb.create_sheet("3-分析-改进项分析单")
render_form(ws3,"改进项分析单（分析阶段）",
    "关联：主诉求单号    接纳后填写闭环单号(校验合法性)进入闭环;不接纳则打回评审",[
    {"title":"一、单据信息","fields":[
        ("TWO",("分析单号",None,SYS),("关联主诉求单号",None,SYS)),
        ("TWO",("分析日期",None,SYS),("责任人(工号+姓名)",None,SYS)),
        ("SINGLE","当前状态",22,OPT["状态"],SYS),
    ]},
    {"title":"二、分析与接纳","fields":[
        ("TWO","接纳版本",("是否接纳",OPT["是否"])),
        ("TWO",("闭环方法",OPT["闭环方法"]),"计划完成期限"),
        ("SINGLE","不接纳理由（是否接纳为『否』时必填,否则打回评审）",40),
    ]},
    {"title":"三、闭环单号（接纳后必填,需校验合法性）","fields":[
        ("SINGLE","问题单号/需求单号",26),
    ]},
    {"title":"四、分析进展子项（可逐条单独提交,作为外部审视进展的依据）","fields":[
        ("TABLE",["序号","进展说明","提交时间","提交人"],
         [(1,1),(2,4),(5,5),(6,6)],6,[SYS,USER,SYS,SYS],[None,None,None,None]),
    ]},
])


# ---------------- Sheet 4-1: 问题单闭环 ----------------
ws4a=wb.create_sheet("4-1闭环-问题单闭环单")
render_form(ws4a,"问题单闭环单（闭环 · 问题单路径）",
    "关联：主诉求单号 + 分析单号    关注当前进展与闭环效果自测",[
    {"title":"一、单据信息","fields":[
        ("TWO",("问题单号",None,SYS),("关联主诉求单号",None,SYS)),
        ("TWO",("关联分析单号",None,SYS),("当前状态",OPT["状态"],SYS)),
        ("TWO",("闭环类型",None,SYS),("责任人(工号+姓名)",None,SYS)),
        ("SINGLE","SLA时间（闭环完成时限,由闭环人填写可改）",24),
    ]},
    {"title":"二、进展与自测","fields":[
        ("SINGLE","当前进展",50),
        ("SINGLE","闭环效果自测（富文本,可粘贴图片）",60),
    ]},
    {"title":"三、进展子项（可逐条单独提交,作为外部审视进展的依据）","fields":[
        ("TABLE",["序号","进展说明","提交时间","提交人"],
         [(1,1),(2,4),(5,5),(6,6)],6,[SYS,USER,SYS,SYS],[None,None,None,None]),
    ]},
])


# ---------------- Sheet 4-2: 需求闭环 ----------------
ws4b=wb.create_sheet("4-2闭环-需求闭环单")
render_form(ws4b,"需求闭环单（闭环 · 需求路径）",
    "关联：主诉求单号 + 分析单号    关注当前进展与闭环效果自测",[
    {"title":"一、单据信息","fields":[
        ("TWO",("需求单号",None,SYS),("关联主诉求单号",None,SYS)),
        ("TWO",("关联分析单号",None,SYS),("当前状态",OPT["状态"],SYS)),
        ("TWO",("闭环类型",None,SYS),("责任人(工号+姓名)",None,SYS)),
        ("SINGLE","SLA时间（闭环完成时限,由闭环人填写可改）",24),
    ]},
    {"title":"二、进展与自测","fields":[
        ("SINGLE","当前进展",50),
        ("SINGLE","闭环效果自测（富文本,可粘贴图片）",60),
    ]},
    {"title":"三、进展子项（可逐条单独提交,作为外部审视进展的依据）","fields":[
        ("TABLE",["序号","进展说明","提交时间","提交人"],
         [(1,1),(2,4),(5,5),(6,6)],6,[SYS,USER,SYS,SYS],[None,None,None,None]),
    ]},
])


# ---------------- Sheet 5: 验收（谁提出谁验收,不可转单） ----------------
ws5=wb.create_sheet("5-验收-验收单")
render_form(ws5,"验收单（谁提出谁验收,不可转单）",
    "关联：主诉求单号 + 闭环单号    验收人=改进提出人,系统生成,锁定不可转单",[
    {"title":"一、单据信息","fields":[
        ("TWO",("验收单号",None,SYS),("关联主诉求单号",None,SYS)),
        ("TWO",("关联闭环单号",None,SYS),("验收人(工号+姓名)",None,SYS)),
        ("SINGLE","当前状态",22,OPT["状态"],SYS),
    ]},
    {"title":"二、验收结论","fields":[
        ("SINGLE","验收结论（富文本,可粘贴图片）",60),
        ("SINGLE","验收是否通过",24,OPT["验收是否通过"]),
    ]},
])


# ---------------- Sheet 6: 填写约定与编号规则 ----------------
ws6=wb.create_sheet("6-填写约定与编号规则")
ws6.sheet_view.showGridlines=False
for col,w in {"A":22,"B":18,"C":62}.items():
    ws6.column_dimensions[col].width=w
merge_val(ws6,1,1,3,"填写约定与编号规则",FILL_TITLE,F_TITLE,C)
ws6.row_dimensions[1].height=30

merge_val(ws6,2,1,3,"一、填写类型（输入区底色）",FILL_SECTION,F_SECTION,L); box(ws6,2,1,2,3); ws6.row_dimensions[2].height=20
legend=[
 ("用户填写","白底","由用户录入(必填/选填/条件必填)"),
 ("系统生成","绿底(带『系统生成』)","系统自动生成/继承,用户免填(编号/状态/关联/日期/责任人继承/提交人等)"),
 ("预填可改","紫底(带『预填·可改』)","系统预填默认值,用户可修改"),
]
r=3
for a,b,cc in legend:
    ws6.cell(r,1,a).font=F_LABEL
    cell=ws6.cell(r,2,b)
    cell.fill={"白底":FILL_INPUT,"绿底(带『系统生成』)":FILL_SYS,"紫底(带『预填·可改』)":FILL_PRE}[b]
    cell.font=F_HINT if b!="白底" else F_VALUE; cell.alignment=C
    ws6.cell(r,3,cc).font=F_VALUE
    for i in range(1,4): ws6.cell(r,i).border=BORDER; ws6.cell(r,i).alignment=L if i==3 else C
    ws6.row_dimensions[r].height=26; r+=1

r+=1
merge_val(ws6,r,1,3,"二、编号规则",FILL_SECTION,F_SECTION,L); box(ws6,r,1,r,3); ws6.row_dimensions[r].height=20; r+=1
for i,h in enumerate(["编号","格式","说明"],start=1):
    cell=ws6.cell(r,i,h); cell.fill=FILL_TBLH; cell.font=F_TBLH; cell.alignment=C; cell.border=BORDER
ws6.row_dimensions[r].height=22; r+=1
ids=[
 ("主诉求单号","QI-YYYY-NNN","主单唯一编号,全流程主键"),
 ("分析单号","AN-YYYY-NNN","改进项分析单编号(系统生成)"),
 ("问题单号","PC-YYYY-NNN","问题单闭环单编号(系统生成)"),
 ("需求单号","RC-YYYY-NNN","需求闭环单编号(系统生成)"),
 ("验收单号","AC-YYYY-NNN","验收单编号(系统生成)"),
 ("关联yw系统单号","YW-YYYY-NNN","提出单关联的业务系统原始单号(用户填写)"),
]
for a,b,cc in ids:
    ws6.cell(r,1,a).font=F_LABEL; ws6.cell(r,1,a).fill=FILL_LABEL
    ws6.cell(r,2,b).font=F_VALUE; ws6.cell(r,3,cc).font=F_VALUE
    for i in range(1,4): ws6.cell(r,i).border=BORDER; ws6.cell(r,i).alignment=C if i==2 else L
    ws6.row_dimensions[r].height=24; r+=1
r+=1
merge_val(ws6,r,1,3,"三、其他约定：审批人为固定下拉名单(系统配置);闭环自测/验收结论支持富文本与粘贴图片;"
                   "责任人 评审填写→分析继承→闭环继承;验收人=提出人,锁定不可转单。",
          FILL_SUB,F_NOTE,L); ws6.row_dimensions[r].height=46


# ---------------- 调整 Sheet 顺序 ----------------
order=["0-流程总览","1-提出-质量改进诉求单","1.5-改进诉求台账(中枢)","2-评审-评审记录单",
       "3-分析-改进项分析单","4-1闭环-问题单闭环单","4-2闭环-需求闭环单",
       "5-验收-验收单","6-填写约定与编号规则"]
wb._sheets.sort(key=lambda s: order.index(s.title))

out=os.path.join(os.path.dirname(os.path.abspath(__file__)),"质量改进诉求流程表单.xlsx")
wb.save(out)
print("saved:",out)
print("sheets:",wb.sheetnames)
